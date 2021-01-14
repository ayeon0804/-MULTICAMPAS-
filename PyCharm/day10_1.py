# [엑셀 파일 다루기]
# 관련 모듈 : openpyxl

# 엑셀 구성요소
# 워크북(workbook=엑셀파일) > 워크시트(worksheet) > 셀(Cell)

from openpyxl import *
import os

# 엑셀파일(xlsx, xls) 불러오기
# 엑셀객체 = load_workbook(url, data_only=True)

print(os.getcwd())
wb = load_workbook('data/sample.xlsx', data_only=True)
print(wb)
print(type(wb))

# 워크시트 객체생성
# 워크시트객체변수 = 워크북객체[워크시트명]

ws = wb['영업사원매출']
print(ws, type(ws))
# <Worksheet "영업사원매출"> <class 'openpyxl.worksheet.worksheet.Worksheet'>

# 워크시트안의 셀 주소
# 워크시트객체['컬럼명행번호'].value
# 워크시트객체.cell(x,y).value
# x, y는 행과 열의 인덱싱으로 1부터 시작

print(ws['A1'].value)  # Sap Co.
print(ws.cell(1, 1).value)  # Sap Co.
print(ws['B2'].value)  # 경기수원대리점
print(ws.cell(2, 2).value)  # 경기수원대리점

# for문을 이용해서 특정 셀안의 내용 출력
print('=' * 50)
for row in ws['A1':'G1']:
    for cell in row:
        print(cell.value, end=' ')
print('\n')
# Sap Co. 대리점 영업사원 전월 금월 TEAM 총 판매수량

# ws.rows 모든행
# ws.columns 모든 열
print('===== 모든행 =======')
for row in ws.rows:
    print(row)

print('===== 모든컬럼 =======')
for column in ws.columns:
    print(column)

print('===== 모든행의 셀안의 값 =======')
for row in ws.rows:
    for cell in row:
        print(cell.value, end = ' ')
    print()

print('===== 모든컬럼의 셀안의 값 =======')
for column in ws.columns:
    for cell in column:
        print(cell.value, end = ' ')
    print()
    print('='*50)

print('===== 모든 셀안의 값을 2차원 리스트로 저장(행방향) =======')
all_values = []
for row in ws.rows:
    temp = []
    for cell in row:
        temp.append(cell.value)
    all_values.append(temp)
print(all_values)

print(all_values[0])
print('='*100)
for a, b, c, d, e, f, g in all_values[1:]:
    print(a, b, c, d, e, f, g)

# 금월 매출액이 가장 높은 영업 사원은?
# 금월 매출액의 최대값 구하기

result_list = []
for a, b, c, d, e, f, g in all_values[1:]:
    result_list.append(e)

print(f'금월 매출액의 최대값 = {max(result_list)}')

for a, b, c, d, e, f, g in all_values[1:]:
    if e == max(result_list):
        print(f'금월 매출액이 가장 높은 영없 사원은 {b} {f}팀 {c}사원입니다.')

# 금월 영업 실적이 전월 영업실적보다 떨어진 영업 사원 목록은?
# 금월영업실적 - 전월영업실적 < 0

for a, b, c, d, e, f, g in all_values[1:]:
    if e - d < 0:
        print(b, c, e-d)

# 퀴즈 : 경기도 지역의 영업 사원 목록 테이타만 csv 파일로 저장하기

gg_list = []
for a, b, c, d, e, f, g in all_values[1:]:
    if b.find('경기') != -1:
        gg_list.append([a, b, c, d, e, f, g])

gg_list.insert(0, all_values[0])
print(gg_list)

import pandas as pd

df = pd.DataFrame(gg_list[1:], columns = gg_list[0])
df.to_csv('output/sample_result.csv', encoding='utf-8', index = False)

# 워크북 객체 닫기
wb.close()
